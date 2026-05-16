"""
客户数据导入脚本

从 CSV 或 XLSX 文件导入客户数据到数据库。

使用方式：
    # 导入单个文件（csv 或 xlsx）
    .venv/Scripts/python.exe scripts/import_customers.py doc/customer2.xlsx

    # 导入多个文件
    .venv/Scripts/python.exe scripts/import_customers.py doc/customer2.xlsx doc/customer3.xlsx

    # 无参数时导入默认文件
    .venv/Scripts/python.exe scripts/import_customers.py

功能：
- 自动映射列名到数据库字段名
- 支持 CSV 和 XLSX 两种文件格式
- 处理日期格式转换
- 处理空值（空字符串转为 None）
- 批量插入到 customers 表
- 跳过表头行和字段说明行
- 支持断点续导（跳过已存在的 original_id 记录）
"""

import csv
import sys
from datetime import datetime
from pathlib import Path

# 将项目根目录加入 sys.path，以便导入 app 模块
PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

from sqlalchemy import select
from sqlalchemy.orm import Session

from app.core.db import get_session_factory
from app.models.customer import Customer

# 默认导入文件列表（无命令行参数时使用）
DEFAULT_FILES = [
    PROJECT_ROOT / "doc" / "customer.csv",
    PROJECT_ROOT / "doc" / "customer.xlsx",
]

# CSV/XLSX 列名 → 数据库字段名的映射
COLUMN_MAPPING: dict[str, str] = {
    "姓名": "name",
    "联系电话": "phone",
    "微信": "wechat",
    "微信状态": "wechat_status",
    "QQ": "qq",
    "省份": "province",
    "地域": "region",
    "年级": "grade",
    "意向度": "intention",
    "反馈状态": "feedback_status",
    "客户阶段": "customer_stage",
    "来源名称": "source_name",
    "归属人": "owner",
    "一级项目": "primary_project",
    "项目": "project",
    "事业部": "business_dept",
    "呼叫部": "call_dept",
    "呼叫组": "call_group",
    "广告商": "advertiser",
    "着陆页": "landing_page",
    "分配方式": "assign_method",
    "分配类型": "assign_type",
    "分配时间": "assigned_at",
    "创建人": "creator",
    "创建人归属机构": "creator_org",
    "首次咨询师": "first_consultant",
    "最后咨询师": "last_consultant",
    "首次分配归属机构": "first_assign_org",
    "首次分配归属人": "first_assign_person",
    "首次分配时间": "first_assign_time",
    "最后一次首咨分配时间": "last_first_consult_time",
    "最后首咨分配归属人": "last_first_consult_person",
    "报名次数": "registration_count",
    "当日外呼次数": "daily_outbound_count",
    "当日呼通的次数": "daily_connected_count",
    "当日接通时长": "daily_connected_duration",
    "Ip": "ip",
    "Ip省份": "ip_province",
    "Ip城市": "ip_city",
    "备注": "remark",
    "标签": "tag",
    "无格式聊天记录": "raw_chat_records",
    "聊天记录": "chat_records",
    "客户ID": "original_id",
}

# 日期时间字段列表，需要做格式转换
DATETIME_FIELDS = {
    "assigned_at",
    "first_assign_time",
    "last_first_consult_time",
}

# 整数字段列表，需要做 int 转换
INTEGER_FIELDS = {
    "registration_count",
    "daily_outbound_count",
    "daily_connected_count",
    "daily_connected_duration",
}


def parse_datetime(value: str) -> datetime | None:
    """
    解析日期时间字符串

    支持常见格式：YYYY-MM-DD HH:MM:SS、YYYY/MM/DD HH:MM:SS、YYYY-MM-DD 等。
    解析失败返回 None。
    """
    if not value or not value.strip():
        return None

    value = value.strip()

    # 常见的日期时间格式
    formats = [
        "%Y-%m-%d %H:%M:%S",
        "%Y/%m/%d %H:%M:%S",
        "%Y-%m-%d",
        "%Y/%m/%d",
        "%Y-%m-%d %H:%M:%S.%f",
    ]

    for fmt in formats:
        try:
            return datetime.strptime(value, fmt)
        except ValueError:
            continue

    print(f"  警告: 无法解析日期 '{value}'，跳过该字段")
    return None


def parse_int(value: str) -> int | None:
    """
    解析整数字符串

    空值或非数字返回 None。
    """
    if not value or not value.strip():
        return None

    try:
        return int(float(value.strip()))
    except (ValueError, TypeError):
        return None


def clean_value(field_name: str, raw_value: str) -> object:
    """
    清洗单个字段值

    - 空字符串 → None
    - 日期时间字段 → datetime 对象
    - 整数字段 → int
    - 其他字段 → 原始字符串
    """
    if not raw_value or not raw_value.strip():
        return None

    if field_name in DATETIME_FIELDS:
        return parse_datetime(raw_value)

    if field_name in INTEGER_FIELDS:
        return parse_int(raw_value)

    return raw_value.strip()


def row_to_customer_data(
    row: dict[str, str],
    header_map: dict[int, str],
) -> dict[str, object]:
    """
    将行数据转换为 Customer 模型可接受的字段字典

    Args:
        row: 行数据（列名→值的字典）
        header_map: 列名→数据库字段名的映射

    Returns:
        字段名字典，可直接传给 Customer(**data)
    """
    data: dict[str, object] = {}

    for csv_col, db_field in header_map.items():
        raw_value = row.get(csv_col, "")
        value = clean_value(db_field, raw_value)
        if value is not None:
            data[db_field] = value

    return data


def read_xlsx(path: Path) -> list[dict[str, str]]:
    """
    读取 XLSX 文件，返回行列表（每行为 {列名: 值} 字典）

    所有单元格值转为字符串，与 CSV DictReader 格式保持一致。

    Args:
        path: XLSX 文件路径

    Returns:
        行数据列表，首行为表头已跳过
    """
    from openpyxl import load_workbook

    wb = load_workbook(filename=str(path), data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return []

    # 首行为表头
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    result: list[dict[str, str]] = []

    for row in rows[1:]:
        row_dict: dict[str, str] = {}
        for i, cell in enumerate(row):
            if i < len(headers) and headers[i]:
                # 将所有值转为字符串，空值保持空字符串
                row_dict[headers[i]] = str(cell) if cell is not None else ""
        result.append(row_dict)

    return result


def read_csv(path: Path) -> list[dict[str, str]]:
    """
    读取 CSV 文件，返回行列表（每行为 {列名: 值} 字典）

    Args:
        path: CSV 文件路径

    Returns:
        行数据列表
    """
    result: list[dict[str, str]] = []
    with open(path, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            result.append(dict(row))
    return result


def import_file(file_path: Path) -> None:
    """
    导入单个文件到数据库

    根据文件后缀自动选择 CSV 或 XLSX 读取方式。

    Args:
        file_path: 文件路径
    """
    if not file_path.exists():
        print(f"错误: 文件不存在: {file_path}")
        return

    suffix = file_path.suffix.lower()

    if suffix == ".xlsx":
        print(f"\n读取 XLSX 文件: {file_path}")
        rows = read_xlsx(file_path)
    elif suffix == ".csv":
        print(f"\n读取 CSV 文件: {file_path}")
        rows = read_csv(file_path)
    else:
        print(f"错误: 不支持的文件格式: {suffix}（仅支持 .csv 和 .xlsx）")
        return

    if not rows:
        print("  文件为空，跳过")
        return

    SessionFactory = get_session_factory()
    db: Session = SessionFactory()

    try:
        # 读取已导入的 original_id 集合，用于跳过重复记录
        existing_ids = set(
            db.execute(
                select(Customer.original_id).where(Customer.original_id.isnot(None))
            )
            .scalars()
            .all()
        )

        # 获取列头并构建映射
        first_row = rows[0]
        header_map: dict[str, str] = {}
        unknown_cols: list[str] = []

        for col in first_row.keys():
            if col in COLUMN_MAPPING:
                header_map[col] = COLUMN_MAPPING[col]
            elif col.strip():
                unknown_cols.append(col)

        if unknown_cols:
            print(f"  提示: 以下列未映射，将被忽略: {unknown_cols}")

        imported_count = 0
        skipped_count = 0
        error_count = 0

        for row in rows:
            # 跳过空行（所有值都为空字符串或 None）
            if not any(v for v in row.values() if v):
                continue

            data = row_to_customer_data(row, header_map)

            # 如果没有有效数据，跳过该行
            if not data:
                continue

            # 通过 original_id 去重，跳过已导入的记录
            original_id = data.get("original_id")
            if original_id and original_id in existing_ids:
                skipped_count += 1
                continue

            try:
                customer = Customer(**data)
                db.add(customer)
                imported_count += 1

                # 每处理 100 条提交一次，减少内存占用
                if imported_count % 100 == 0:
                    db.commit()
                    print(f"  已导入 {imported_count} 条...")
            except Exception as e:
                error_count += 1
                print(f"  导入失败: {e}")
                db.rollback()

        # 提交剩余记录
        db.commit()

        print(
            f"  导入完成: 成功 {imported_count} 条, 跳过 {skipped_count} 条, 失败 {error_count} 条"
        )

    except Exception as e:
        db.rollback()
        print(f"  导入失败: {e}")
        raise
    finally:
        db.close()


def main() -> None:
    """主入口：解析命令行参数，依次导入文件"""
    if len(sys.argv) > 1:
        # 命令行指定了文件
        files = [Path(arg) for arg in sys.argv[1:]]
    else:
        # 无参数时使用默认文件列表
        files = DEFAULT_FILES

    print(f"准备导入 {len(files)} 个文件")

    for f in files:
        import_file(f)

    print("\n全部文件导入完毕")


if __name__ == "__main__":
    main()
